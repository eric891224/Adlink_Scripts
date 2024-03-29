import os
import sys
import logging

from enum import Enum
from typing import List, Dict, Tuple
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

class AutoTestType(Enum):
    AVAILABILITY = 'A'
    FUNCTIONAL = 'F'

class CommandStatus(Enum):
    AVAILABLE = 'A'
    UNAVAILABLE = 'U'

class NetFn(Enum):
    CHASSIS = '0x00'
    BRIDGE = '0x02'
    SENSOR_EVENT = '0x04'
    APPLICATION = '0x06'
    FIRMWARE = '0x08'
    STORAGE = '0x0a'
    TRANSPORT = '0x0c'
    PICMG = '0x2c'

class VerificationType(Enum):
    ACCURACY = 'accuracy'
    BEHAVIOR = 'behavior'

class Result(Enum):
    PASS = 'P'
    FAIL = 'F'

class PassLevel(Enum):
    ALL_MATCH = 'A'
    PARTIAL_MATCH = 'P'
    IGNORED = 'I'

class DotDict(dict):
    """
        dot.notation access to dictionary attributes
    """      
    def __getattr__(*args):
        val = dict.get(*args)         
        return DotDict(val) if type(val) is dict else val  
        
    __setattr__ = dict.__setitem__     
    __delattr__ = dict.__delitem__ 

GREEN_FILL = PatternFill(start_color='00ff00', fill_type='solid')
DARK_GREEN_FILL = PatternFill(start_color='008000', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='ffff00', fill_type='solid')
RED_FILL = PatternFill(start_color='ff0000', fill_type='solid')

LOGFILE_NAME = "logfile.log"

SENSOR_THRESHOLD = {
    "Lower Non-Critical": "LNC",
    "Lower Critical": "LC",
    "Lower Non-Recoverable": "LNR",
    "Upper Non-Critical": "UNC",
    "Upper Critical": "UC",
    "Upper Non-Recoverable": "UNR"
}

def parseNetFn(netfn: str) -> NetFn:
    if netfn == 'Chassis': return NetFn.CHASSIS
    elif netfn == 'Bridge': return NetFn.BRIDGE
    elif netfn == 'S/E': return NetFn.SENSOR_EVENT
    elif netfn == 'App': return NetFn.APPLICATION
    elif netfn == 'Firmware': return NetFn.FIRMWARE
    elif netfn == 'Storage': return NetFn.STORAGE
    elif netfn == 'Transport': return NetFn.TRANSPORT
    elif netfn == 'PICMG': return NetFn.PICMG
    else: raise Exception(f"Netfn '{netfn}' not found.")

def parseCmd(cmd: str) -> str:
    return f'0x{cmd.rstrip("h")}'

def parseVerificationType(verificationType: str) -> str:
    if verificationType == 'A': return VerificationType.ACCURACY
    elif verificationType == 'B': return VerificationType.BEHAVIOR
    else: raise Exception(f"Verification Type {verificationType} not defined")

def parseRawFunctionName(functionName: str) -> str:
    return functionName.replace(' ', '').replace('-', '_')

# deprecated
def parseLabelKey(key: str) -> Tuple[str]:
    keys = key.lstrip('[').rstrip(']').replace(' ', "").split(',')
    return tuple(keys)

def parseFruInfo(stdout: str) -> Dict[str, str]:
    stdout = stdout.split('\n')
    fruInfo = {}
    for line in stdout:
        if ':' in line:
            line = line.split(':')
            fruInfo[line[0].strip(" ")] = line[1].strip(' ') if len(line) <= 2 else ':'.join(line[1:]).strip(' ')

    return fruInfo

def parseSensorInfo(stdout: str) -> Dict[str, str]:
    stdout = stdout.split('\n')
    sensorInfo = {}
    for line in stdout:
        if ':' in line:
            line = line.split(':')
            sensorInfo[line[0].strip(' ')] = line[1].strip(' ') if len(line) <= 2 else ':'.join(line[1:]).strip(' ')

    return sensorInfo

def isWorkSheetColEmpty(workSheet: Worksheet, startRow: int, colNum: int) -> bool:
    for row in workSheet.iter_rows(min_row=startRow, max_row=workSheet.max_row, min_col=colNum, max_col=colNum):
        for cell in row:
            if cell.value is not None:  # Check if cell is not empty
                return False
    return True

def getAccuracyMetric(accuracy: int) -> Tuple[PassLevel, PatternFill]:
    if accuracy == 100:
        return PassLevel.ALL_MATCH, GREEN_FILL
    elif accuracy > 60:
        return PassLevel.PARTIAL_MATCH, YELLOW_FILL
    else:
        return None, RED_FILL

def getLoggingFileHandler(outputPath: str) -> List[logging.FileHandler]:
    fileHandler = logging.FileHandler(os.path.join(outputPath, LOGFILE_NAME), 'w')
    fileHandler.setLevel(logging.DEBUG)

    streamHandler = logging.StreamHandler(stream=sys.stdout)
    streamHandler.setLevel(logging.INFO)

    return [fileHandler, streamHandler]