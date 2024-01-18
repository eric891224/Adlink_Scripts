import os
from openpyxl.styles import PatternFill

PROJECTS_ROOT = os.path.abspath('./projects/')

GREEN_FILL = PatternFill(start_color='00ff00', fill_type='solid')
DARK_GREEN_FILL = PatternFill(start_color='008000', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='ffff00', fill_type='solid')
RED_FILL = PatternFill(start_color='ff0000', fill_type='solid')

LOGFILE_NAME = "logfile.log"

SENSOR_THRESHOLD = {
    "Lower Non-Critical": "LNC",
    "Lower Critical": "LC",
    "Lower Non-Recoverable": "LNR",
    "Upper Non-Critical": "UNC",
    "Upper Critical": "UC",
    "Upper Non-Recoverable": "UNR"
}