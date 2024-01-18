import os
import sys
import logging

from typing import List, Tuple
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .dotDict import DotDict
from .enums import PassLevel
from .globalVars import PROJECTS_ROOT, GREEN_FILL, YELLOW_FILL, RED_FILL, LOGFILE_NAME

def getProjectDir(projectName: str) -> str:
    return os.path.join(PROJECTS_ROOT, projectName)

def getInputFilePath(projectName: str) -> str:
    return os.path.join(getProjectDir(projectName), f'{projectName}.xlsx')

def getLabelsDir(projectName: str) -> str:
    return os.path.join(getProjectDir(projectName), 'labels')

def getOutputDir(outputRoot: str, projectName: str) -> str:
    return os.path.join(outputRoot, projectName)

def isWorkSheetColEmpty(workSheet: Worksheet, startRow: int, colNum: int) -> bool:
    for row in workSheet.iter_rows(min_row=startRow, max_row=workSheet.max_row, min_col=colNum, max_col=colNum):
        for cell in row:
            if cell.value is not None:  # Check if cell is not empty
                return False
    return True

def getAccuracyMetric(accuracy: int) -> PassLevel:
    if accuracy == 100:
        return PassLevel.ALL_MATCH, GREEN_FILL
    elif accuracy > 60:
        return PassLevel.PARTIAL_MATCH, YELLOW_FILL
    else:
        return ""

def getLoggingFileHandler(outputPath: str) -> List[logging.FileHandler]:
    fileHandler = logging.FileHandler(os.path.join(outputPath, LOGFILE_NAME), 'w')
    fileHandler.setLevel(logging.DEBUG)

    streamHandler = logging.StreamHandler(stream=sys.stdout)
    streamHandler.setLevel(logging.INFO)

    return [fileHandler, streamHandler]

def getProjectConfigLogs(projectConfig: DotDict) -> str:
    logs = f"Project: {projectConfig.projectName}\nPass Level: {projectConfig.passLevel}"
    
    if projectConfig.tester:
        logs += f"Tester: {projectConfig.tester}\n"
    if projectConfig.hardwareVersion:
        logs += f"HW Version: {projectConfig.hardwareVersion}\n"
    if projectConfig.softwareVersion:
        logs += f"SW Version: {projectConfig.softwareVersion}\n"

    return logs