import os
import json
import subprocess
import logging

from importlib import import_module
from datetime import datetime
from typing import Tuple, Dict, Union

from tqdm import tqdm
from pandas import ExcelFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from defs.dotDict import DotDict
from defs.globalVars import GREEN_FILL, DARK_GREEN_FILL, YELLOW_FILL, RED_FILL, LOGFILE_NAME
from defs.enums import AutoTestType, CommandStatus, Result, PassLevel, VerificationType
from defs.parsers import parseNetFn, parseCmd, parseVerificationType, parseRawFunctionName
from defs.functions import isWorkSheetColEmpty, getAccuracyMetric, getProjectConfigLogs

class IPMIAutoTest:
    def __init__(
        self, 
        excelFile: ExcelFile, 
        workBook: Workbook, 
        projectConfig: DotDict,
        labelsDir: str,
        outputDir: str, 
        isOutOfBand: bool=False,
        **kwargs: Dict[str, bool]    # doRawAvailabilityTest, doRawFunctionalTest, doFruTest, doSensorTest
    ) -> None:
        self.excelFile = excelFile
        self.workBook = workBook
        self.projectConfig = projectConfig
        self.labelsDir = os.path.abspath(labelsDir)
        self.outputDir = os.path.abspath(outputDir)
        self.isOutOfBand = isOutOfBand
        self.tasks = DotDict(kwargs)

        self.time = datetime.now().strftime("%Y-%b-%d_%H-%M")
        self.logger = logging.getLogger(f'main.{self.__str__()}')

        if isOutOfBand:
            self.__commandTemplate = f"ipmitool -C {self.projectConfig.cypherSuite} -I lanplus -H {self.projectConfig.ip} -U {self.projectConfig.userName} -P {self.projectConfig.password}"
        else:
            self.__commandTemplate = f"ipmitool{' -I wmi' if os.name == 'nt' else ''}"

    def __str__(self) -> str:
        return 'IPMIAutoCommandTest'
    
    def __checkConnection(self) -> bool:
        ping = subprocess.Popen(
            f"ping {self.projectConfig.ip} {'-n' if os.name == 'nt' else '-c'} 4", 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            universal_newlines=True,
            shell=True
        )

        stdout, stderr = ping.communicate()

        if "received, 100% packet loss" in stdout or "received, 100% packet loss" in stderr:
            return False

        return True
    
    def __checkIPMIToolViability(self) -> Tuple[bool,str]:
        if self.isOutOfBand and not self.__checkConnection():
            self.logger.error('PING FAIL')
            return False, f"Unable to ping {self.projectConfig.ip}, make sure the connection is viable."
            
        _, stderr = self.__rawCommand()
        if 'Could not open device' in stderr:
            self.logger.error('IPMITOOL FAIL')
            return False, "Cannot access ipmitool, make sure the tool is properly installed."
            
        return True, None
    
    def __generalCommand(
        self, 
        cmdType: str="", 
        *args: Tuple[str]
    ) -> Tuple[str, str]:
        shell = f"{self.__commandTemplate} {cmdType}".rstrip(" ")
        for arg in args:
            shell += f' {arg}'

        res = subprocess.Popen(
            shell,
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            universal_newlines=True,
            shell=True
        )

        return res.communicate()

    def __rawCommand(
        self, 
        netfn: str="", 
        cmd: str="", 
        *args: Tuple[str]
    ) -> Tuple[str, str]:
        return self.__generalCommand("raw", netfn, cmd, *args)

    def __writeCell(
        self, 
        workSheet: Worksheet,
        row: int, 
        col: int, 
        value: str="", 
        fill: PatternFill=None
    ) -> None:
        cell = workSheet.cell(row=row, column=col)
        cell.value = value
        if bool(fill): 
            cell.fill = fill

    def __testSingleRawFunction(
        self, 
        functionName: str, 
        netfn: str, 
        cmd: str, 
        numData: int, 
        needVerify: bool=False,
        verificationType: str=None
    ) -> Tuple[Union[int, Result], str]:
        labels = None
        if verificationType == VerificationType.ACCURACY:
            labels = json.load(open(os.path.join(self.labelsDir, functionName[0].lower() + functionName[1:] + 'Labels.json'))).data
        # labels = {parseLabelKey(key): value for key, value in labels.items()}
        # module = globals()[f'test{functionName}']
        module = import_module(f'projects.{self.projectConfig.projectName}.raw.test{functionName}')
        testCase = getattr(module, f'Test{functionName}')(
            self,
            netfn,
            cmd,
            numData,
            needVerify=needVerify,
            labels=labels,
            verificationType=verificationType
        )
        result, info = testCase.test()

        return result, info

    def runTest(self) -> None:
        self.logger.info(getProjectConfigLogs(self.projectConfig))

        viable, res = self.__checkIPMIToolViability()
        if not viable: raise Exception(res)

        if self.tasks.doRawAvailabilityTest:
            self.testRawAvailability()
        if self.tasks.doRawFunctionalTest:
            self.testRawFunction()
        # if self.tasks.doFruTest:
        #     self.testFru()
        # if self.tasks.doSensorTest:
        #     self.testSensor()

        self.saveOutput()

    def saveOutput(self) -> None:   # TODO
        outputPath = os.path.join(self.outputDir, self.time)
        if not os.path.exists(outputPath):
            os.makedirs(outputPath)

        mode = 'OOB' if self.isOutOfBand else 'IB'
        fileName = '_'.join([self.projectConfig.projectName, mode]) + '.xlsx'

        self.workBook.save(os.path.join(outputPath,  fileName))
        self.logger.info(f"Result generated at {os.path.join(outputPath, fileName)}")
        logging.shutdown()
        os.rename(
            os.path.join(os.path.dirname(os.path.realpath(__file__)), LOGFILE_NAME),
            os.path.join(outputPath, LOGFILE_NAME)
        )

    def testRawAvailability(self):
        self.logger.info("===Testing Raw CMD Availability===")
        rawCmdDf = self.excelFile.parse('Raw CMD', header=1)
        rawCmdSheet = self.workBook['Raw CMD']

        for row in rawCmdSheet.iter_rows(max_row=2, values_only=True):
            labels = row
        
        supColNum = labels.index('Availability (A: available/U: unavailable)') + 1
        resColNum = labels.index('Error Response') + 1

        for functionName, netfn, cmd, isNan, rowNum in tqdm(
            zip(
                rawCmdDf['Function Name'],
                rawCmdDf['NetFn'], 
                rawCmdDf['CMD'], 
                rawCmdDf['NetFn'].isnull(), 
                range(3, len(rawCmdDf['NetFn'])+3)
            ),
            desc='Testing...',
            total=len(rawCmdDf['NetFn']),
            ncols=100,
            leave=True
        ):
            if isNan:
                continue
            if functionName == 'Warm Reset' or functionName == 'Cold Reset':    # TODO: handle rest function
                continue
            if 'SOL' in functionName:    # TODO: handle sol
                continue

            stdout, stderr = self.__rawCommand(
                parseNetFn(netfn).value, 
                parseCmd(cmd),
            )
            self.logger.debug(stdout.rstrip("\n") if stdout else stderr.rstrip("\n"))

            if 'Invalid command' in stderr or 'Unknown' in stderr:
                self.__writeCell(rawCmdSheet, rowNum, supColNum, CommandStatus.UNAVAILABLE.value, RED_FILL)
                self.__writeCell(rawCmdSheet, rowNum, resColNum, stderr)
            else:
                self.__writeCell(rawCmdSheet, rowNum, supColNum, CommandStatus.AVAILABLE.value, GREEN_FILL)

    def testRawFunction(self):
        rawCmdDf = self.excelFile.parse('Raw CMD', header=1)
        rawCmdSheet = self.workBook['Raw CMD']

        for row in rawCmdSheet.iter_rows(max_row=2, values_only=True):
            labels = row

        cmdStatusColNum = labels.index('Availability (A: available/U: unavailable)') + 1
        resultInfoColNum = labels.index('Info') + 1
        resultColNum = labels.index('Result (P: pass/F: fail/%: accuracy)') + 1
        passColnum = labels.index('Pass Level (A: all match, P: partial match, I: ignored)') + 1

        if (isWorkSheetColEmpty(rawCmdSheet, 3, cmdStatusColNum)):
            self.logger.info("IPMI availability not tested, proceed to test raw availability...")
            self.testRawAvailability()

        self.logger.info("===Testing Raw Function===")

        for functionName, netfn, cmd, numData, testType, needVerify, verificationType, isNan, rowNum in tqdm(
            zip(
                rawCmdDf['Function Name'],
                rawCmdDf['NetFn'], 
                rawCmdDf['CMD'],
                rawCmdDf['Request Length'],
                rawCmdDf['Test Type (A: availability test/F: functional test)'],
                rawCmdDf['Need Verify (Y: yes/N: no)'],
                rawCmdDf['Verification Type (A: accuracy/B: behavior)'],
                rawCmdDf['NetFn'].isnull(), 
                range(3, len(rawCmdDf['Request Length'])+3)
            ),
            desc='Testing...',
            total=len(rawCmdDf['Request Length']),
            ncols=100,
            leave=True
        ):
            if isNan:
                continue
            if functionName == 'Warm Reset' or functionName == 'Cold Reset':    # TODO: handle rest function
                continue
            if 'SOL' in functionName:    # TODO: handle sol
                continue

            cmdStatus = rawCmdSheet.cell(rowNum, cmdStatusColNum).value

            result = passLevel = resultColor = passLevelColor = None

            if testType == AutoTestType.AVAILABILITY.value:
                if cmdStatus == None:
                    continue
                elif cmdStatus == CommandStatus.UNAVAILABLE.value:
                    result = Result.FAIL
                    passLevel = ""
                    resultColor = passLevelColor = RED_FILL
                else: # available
                    passLevel = PassLevel.IGNORED
                    if passLevel < self.projectConfig.passLevel:
                        result = Result.FAIL
                        resultColor = passLevelColor = RED_FILL
                    else:
                        result = Result.PASS
                        resultColor = GREEN_FILL
                        passLevelColor = DARK_GREEN_FILL
            else: # funcional Test
                result, info = self.__testSingleRawFunction(
                    parseRawFunctionName(functionName),
                    parseNetFn(netfn).value,
                    parseCmd(cmd),
                    int(numData),
                    needVerify=needVerify,
                    verificationType=parseVerificationType(verificationType)
                )

                if info is not None:
                    self.__writeCell(rawCmdSheet, rowNum, resultInfoColNum, info)

                if type(result) == int: # % accuracy
                    passLevel = getAccuracyMetric(result)
                    if passLevel == "":
                        resultColor = passLevelColor = RED_FILL
                    elif passLevel == PassLevel.PARTIAL_MATCH:
                        if passLevel < self.projectConfig.passLevel:
                            resultColor = passLevelColor = RED_FILL
                        else:
                            resultColor = YELLOW_FILL
                            passLevelColor = YELLOW_FILL
                    elif passLevel == PassLevel.ALL_MATCH:
                        resultColor = passLevelColor = GREEN_FILL
                    else: # IGNORED
                        resultColor = GREEN_FILL
                        passLevelColor = DARK_GREEN_FILL
                    result = DotDict({"value": f'{result}%'})
                else:   # Result
                    if result == Result.PASS:
                        passLevel = PassLevel.ALL_MATCH if needVerify else PassLevel.IGNORED
                        resultColor = GREEN_FILL
                        passLevelColor = GREEN_FILL if needVerify else DARK_GREEN_FILL
                    elif result == Result.FAIL:
                        passLevel = ""
                        resultColor = passLevelColor = RED_FILL

            passLevel = DotDict({"value": ""}) if type(passLevel) == str else passLevel

            self.__writeCell(rawCmdSheet, rowNum, resultColNum, result.value, resultColor)
            self.__writeCell(rawCmdSheet, rowNum, passColnum, passLevel.value, passLevelColor)

    def _devTest(self):
        stdout, stderr = self.__generalCommand("power", "status")
        print(stdout.strip(' ').strip('\n').split(' ')[-1])
        statuses = stdout.split('\n')
        for status in statuses:
            if "System Power" in status:
                chassisStatus = status.split(':')[1].strip(' ')
                print(chassisStatus)
                break