import os
import logging
import sys
import signal

# from importlib import resources
from functools import partial
from pandas import ExcelFile
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))

from defs.dotDict import DotDict
from defs.globalVars import LOGFILE_NAME
from defs.parsers import parsePassLevel
from defs.functions import getInputFilePath, getLabelsDir, getOutputDir, getLoggingFileHandler

from autoTest import IPMIAutoTest
from argParser import IPMIAutoTestParser

def exit_gracefully(logger: logging.Logger, sig: int, frame) -> None:
    logger.info('Program terminated')
    logging.shutdown()
    os.remove(os.path.join(os.path.dirname(os.path.realpath(__file__)), LOGFILE_NAME))
    sys.exit(0)

if __name__ == '__main__':
    parser = IPMIAutoTestParser()
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG, 
        format="[%(asctime)s][%(name)-5s][%(levelname)-5s] %(message)s (%(filename)s:%(lineno)d)",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=getLoggingFileHandler(os.path.dirname(os.path.realpath(__file__))),
    )

    signal.signal(signal.SIGINT, partial(exit_gracefully, logging.getLogger('main')))

    excelFile = ExcelFile(getInputFilePath(args.project_name))
    workBook = load_workbook(getInputFilePath(args.project_name))
    # with resources.path(f'project.{args.project_name}', f'{args.project_name}.xlsx') as inputFilePath:
    #     excelFile = ExcelFile(inputFilePath)
    #     workBook = load_workbook(inputFilePath)
    projectConfig = DotDict({
        "projectName": args.project_name,
        "passLevel": parsePassLevel(args.pass_level),
        "ip": args.ip,
        "cypherSuite": args.cypher_suite,
        "userName": args.user_name,
        "password": args.password,
        "hardwareVersion": args.hardware_version,
        "softwareVersion": args.software_version,
        "tester": args.tester
    })

    testRoutine = IPMIAutoTest(
        excelFile, 
        workBook,
        projectConfig = projectConfig,
        labelsDir=getLabelsDir(args.project_name),
        outputDir=getOutputDir(args.output_directory, args.project_name), 
        isOutOfBand=args.is_out_of_band,
        doRawAvailabilityTest=args.raw_availability_test,
        doRawFunctionalTest=args.raw_functional_test,
        doFruTest=args.fru,
        doSensorTest=args.sensor
    )

    testRoutine.runTest()
    # testRoutine._devTest()