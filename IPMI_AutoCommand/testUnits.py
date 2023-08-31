import os
import subprocess
import logging

from pandas import ExcelFile

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from datetime import datetime
from tqdm import tqdm
from typing import Tuple, List, Dict

from defs import *

class IPMIAutoCommandTest:
    def __init__(
        self, 
        excelFile: ExcelFile, 
        workBook: Workbook, 
        outputDir: str='./result/', 
        isOutOfBand: bool=False,
        **kwargs: Dict[str, bool]    # doRawSupportTest, doFunctionalTest
    ) -> None:
        self.excelFile = excelFile
        self.workBook = workBook
        self.outputDir = os.path.abspath(outputDir)
        self.isOutOfBand = isOutOfBand
        self.tasks = DotDict(kwargs)

        self.time = datetime.now().strftime("%Y-%b-%d_%H-%M")
        self.logger = logging.getLogger(f'main.{self.__str__()}')

        projectInfoDf = self.excelFile.parse('Project Info')
        self.projectInfo = { key: value for key, value in zip(projectInfoDf["Variable"], projectInfoDf["Value"]) }

        if isOutOfBand:
            self.__commandTemplate = f"ipmitool -I lanplus -H {self.projectInfo['IP']} -U {self.projectInfo['Account_Name']} -P {self.projectInfo['Account_Password']}"
        else:
            self.__commandTemplate = "ipmitool"
    
    def __str__(self) -> str:
        return 'IPMIAutoCommandTest'

    def __checkConnection(self) -> bool:
        ping = subprocess.Popen(
            f"ping {self.projectInfo['IP']} {'-n' if os.name == 'nt' else '-c'} 4", 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            universal_newlines=True,
            shell=True
        )

        stdout, stderr = ping.communicate()

        if "received, 100% packet loss" in stdout or "received, 100% packet loss" in stderr:
            return False

        return True
    
    def __checkIPMIToolViability(self) -> Tuple[bool,str]:
        if self.isOutOfBand and not self.__checkConnection():
            self.logger.error('PING FAIL')
            return False, f"Unable to ping {self.projectInfo['IP']}, make sure the connection is viable."
            
        _, stderr = self.__rawCommand()
        if 'Could not open device' in stderr:
            self.logger.error('IPMITOOL FAIL')
            return False, "Cannot access ipmitool, make sure the tool is properly installed."
            
        return True, None
    
    def __generalCommand(
        self, 
        cmdType: str="", 
        *args: List[str]
    ) -> Tuple[str, str]:
        shell = f"{self.__commandTemplate} {cmdType}".rstrip(" ")
        for arg in args:
            shell += f' {arg}'

        res = subprocess.Popen(
            shell,
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            universal_newlines=True,
            shell=True
        )

        return res.communicate()

    def __rawCommand(
        self, 
        netfn: str="", 
        cmd: str="", 
        *args: List[str]
    ) -> Tuple[str, str]:
        # shell = f"{self.__commandTemplate} raw {netfn} {cmd}".rstrip(" ")
        # for arg in args:
        #     shell += f' {arg}'

        # res = subprocess.Popen(
        #     shell,
        #     stdout=subprocess.PIPE, 
        #     stderr=subprocess.PIPE, 
        #     universal_newlines=True,
        #     shell=True
        # )

        # return res.communicate()
        return self.__generalCommand("raw", netfn, cmd, *args)

    def __writeCell(
        self, 
        workSheet: Worksheet,
        row: int, 
        col: int, 
        value: str="", 
        fill: PatternFill=None
    ) -> None:
        cell = workSheet.cell(row=row, column=col)
        cell.value = value
        if bool(fill): 
            cell.fill = fill

    def saveOutput(self) -> None:
        outputPath = os.path.join(self.outputDir, self.time)
        if not os.path.exists(outputPath):
            os.makedirs(outputPath)

        fileNameFormats = self.projectInfo["Output_File_Name"].split("+")
        fileName = 'OUTOFBAND' if self.isOutOfBand else 'INBAND'

        for format in fileNameFormats:
            format = format.strip()
            if (format[0] == '$'):
                fileName += '_' + str(self.projectInfo[format[1:]]).replace(' ', '-')
            else:
                fileName += '_' + format.strip('"').strip("'").strip().replace(' ', '-')

        self.workBook.save(os.path.join(outputPath, fileName + '.xlsx'))
        self.logger.info(f"Result generated at {os.path.join(outputPath, fileName + '.xlsx')}")
        logging.shutdown()
        os.rename(
            os.path.join(os.path.dirname(os.path.realpath(__file__)), LOGFILE_NAME),
            os.path.join(outputPath, LOGFILE_NAME)
        )

    def runTest(self) -> None:
        viable, res = self.__checkIPMIToolViability()
        if not viable: raise Exception(res)

        if self.tasks.doRawSupportTest:
            self.testRawSupport()
        if self.tasks.doFunctionalTest:
            self.testRawFunction()
        if self.tasks.doFruTest:
            self.testFru()
        if self.tasks.doSensorTest:
            self.testSensor()

        self.saveOutput()
        
    def testRawSupport(self) -> None:
        self.logger.info("===Testing Raw CMD Support===")
        rawCmdDf = self.excelFile.parse('Raw CMD', header=1)
        rawCmdSheet = self.workBook['Raw CMD']

        for row in rawCmdSheet.iter_rows(max_row=2, values_only=True):
            labels = row
        
        supColNum = labels.index('Availability (A: available/U: unavailable)') + 1
        resColNum = labels.index('Response') + 1

        for netfn, cmd, isNan, rowNum in tqdm(
            zip(rawCmdDf['NetFn'], rawCmdDf['CMD'], rawCmdDf['NetFn'].isnull(), range(3, len(rawCmdDf['NetFn'])+3)),
            desc='Testing...',
            total=len(rawCmdDf['NetFn']),
            ncols=100,
            leave=True
        ):
            if isNan:
                continue

            stdout, stderr = self.__rawCommand(
                parseNetFn(netfn).value, 
                parseCmd(cmd),
            )
            self.logger.debug(stdout.rstrip("\n") if stdout else stderr.rstrip("\n"))

            if 'Invalid command' in stderr or 'Unknown' in stderr:
                self.__writeCell(rawCmdSheet, rowNum, supColNum, CommandStatus.UNAVAILABLE.value, RED_FILL)
                self.__writeCell(rawCmdSheet, rowNum, resColNum, stderr)
            else:
                self.__writeCell(rawCmdSheet, rowNum, supColNum, CommandStatus.AVAILABLE.value, GREEN_FILL)
                
    def testRawFunction(self) -> None:
        self.logger.info("===Testing Raw Function===")
        pass

    def testFru(self) -> None:
        self.logger.info("===Testing Fru===")
        fruInfoDf = self.excelFile.parse('Fru Info', header=0)
        fruInfoSheet = self.workBook['Fru Info']

        for row in fruInfoSheet.iter_rows(max_row=1, values_only=True):
            labels = row
        
        resColNum = labels.index('Response') + 1
        resultColNum = labels.index('Result (P: pass/F: fail)') + 1
        passColNum = labels.index('Pass Level (A: all match, L: length match, I: ignored)') + 1

        stdout, stderr = self.__generalCommand("fru")

        if not stderr:
            self.logger.debug(stdout)
            fruInfo = parseFruInfo(stdout)
            for IPMITag, expectedData, rowNum in tqdm(
                zip(fruInfoDf['IPMI Tag'], fruInfoDf['Expected Data'], range(2, len(fruInfoDf['IPMI Tag'])+2)),
                desc='Testing...',
                total=len(fruInfoDf['IPMI Tag']),
                ncols=100,
                leave=True
            ):
                hasTag = IPMITag in fruInfo.keys()

                if expectedData == '[ignored]':
                    self.__writeCell(fruInfoSheet, rowNum, resColNum, fruInfo[IPMITag] if hasTag else "")
                    self.__writeCell(fruInfoSheet, rowNum, resultColNum, Result.PASS.value, GREEN_FILL)
                    self.__writeCell(fruInfoSheet, rowNum, passColNum, PassLevel.IGNORED.value, DARK_GREEN_FILL)
                    continue
                
                if hasTag and expectedData == fruInfo[IPMITag]:
                    self.__writeCell(fruInfoSheet, rowNum, resColNum, fruInfo[IPMITag])
                    self.__writeCell(fruInfoSheet, rowNum, resultColNum, Result.PASS.value, GREEN_FILL)
                    self.__writeCell(fruInfoSheet, rowNum, passColNum, PassLevel.ALL_MATCH.value, GREEN_FILL)
                elif hasTag and expectedData != fruInfo[IPMITag] and len(expectedData) == len(fruInfo[IPMITag]):
                    self.__writeCell(fruInfoSheet, rowNum, resColNum, fruInfo[IPMITag])
                    self.__writeCell(fruInfoSheet, rowNum, resultColNum, Result.PASS.value, GREEN_FILL)
                    self.__writeCell(fruInfoSheet, rowNum, passColNum, PassLevel.LENGTH_MATCH.value, GREEN_FILL)
                elif not hasTag and expectedData == '[null]':
                    self.__writeCell(fruInfoSheet, rowNum, resultColNum, Result.PASS.value, GREEN_FILL)
                    self.__writeCell(fruInfoSheet, rowNum, passColNum, PassLevel.ALL_MATCH.value, GREEN_FILL)
                else:
                    self.__writeCell(fruInfoSheet, rowNum, resColNum, fruInfo[IPMITag] if hasTag else "")
                    self.__writeCell(fruInfoSheet, rowNum, resultColNum, Result.FAIL.value, RED_FILL)
                    self.__writeCell(fruInfoSheet, rowNum, passColNum, "", RED_FILL)
        else:
            self.logger.error(f'TODO: error handling\n STDERR: \n{stderr}')
            raise NotImplementedError()

        
    def testSensor(self) -> None:
        self.logger.info("===Testing Sensor===")
        sensorInfoDf = self.excelFile.parse('Sensor Info', header=1)
        sensorInfoSheet = self.workBook['Sensor Info']

        for row in sensorInfoSheet.iter_rows(max_row=2, values_only=True):
            labels = row

        resultColNum = labels.index('Result (P: pass/F: fail)') + 1
        errResColNum = labels.index('Error Response') + 1

        for sensorName, sensorIndex,rowNum in tqdm(
            zip(sensorInfoDf['Sensor Name'], range(len(sensorInfoDf['Sensor Name'])), range(3, len(sensorInfoDf['Sensor Name'])+3)),
            desc='Testing...',
            total=len(sensorInfoDf['Sensor Name']),
            ncols=100,
            leave=True
        ):
            passCount = 0
            stdout, stderr = self.__generalCommand("sensor", "get", sensorName)

            if not stderr:
                if sensorInfoDf['Sensor Type'][sensorIndex].lower() in ['discrete', 'watchdog']:
                    self.__writeCell(sensorInfoSheet, rowNum, resultColNum, Result.PASS.value, GREEN_FILL)
                    continue

                sensorInfo = parseSensorInfo(stdout)
                self.logger.debug(sensorInfo)

                for status in SENSOR_THRESHOLD.keys():
                    statusAbbr = SENSOR_THRESHOLD[status]
                    outputStatusColNum = [colNum for colNum, label in enumerate(labels) if label == statusAbbr][1] + 1
                    if float(sensorInfoDf[statusAbbr][sensorIndex]) == float(sensorInfo[status]):
                        self.__writeCell(sensorInfoSheet, rowNum, outputStatusColNum, sensorInfo[status])
                        passCount += 1
                    else:
                        self.__writeCell(sensorInfoSheet, rowNum, outputStatusColNum, sensorInfo[status], RED_FILL)

                if passCount == 6:
                    self.__writeCell(sensorInfoSheet, rowNum, resultColNum, Result.PASS.value, GREEN_FILL)
                else:
                    self.__writeCell(sensorInfoSheet, rowNum, resultColNum, Result.FAIL.value, RED_FILL)
            else:
                self.logger.debug(stderr.strip('\n'))
                self.__writeCell(sensorInfoSheet, rowNum, resultColNum, Result.FAIL.value, RED_FILL)
                self.__writeCell(sensorInfoSheet, rowNum, errResColNum, stderr)